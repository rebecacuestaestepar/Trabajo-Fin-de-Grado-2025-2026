import openpyxl
import re

DIAS_SEMANA = {
    "LUNES": 1,
    "MARTES": 2,
    "MIÉRCOLES": 3,
    "JUEVES": 4,
    "VIERNES": 5,
    "SÁBADO": 6,
    "DOMINGO": 7,
}


class Clase:
    def __init__(self, asig, cod_asig, grupo, aula, hora_inicio, hora_fin, dia, tipo):
        self.asig = asig
        self.cod_asig = cod_asig
        self.grupo = grupo
        self.aula = aula
        self.hora_inicio = hora_inicio
        self.hora_fin = hora_fin
        self.dia = dia
        self.tipo = tipo
        
    def __repr__(self):
        return f"[{self.dia} | {self.hora_inicio}-{self.hora_fin}] {self.tipo} - {self.asig} (Gr:{self.grupo}) en {self.aula}"


class ConfigurarHorario:
    HORA = re.compile(r"(?i)^\s*hora\s*$")
    TRAMO = re.compile(r"^\s*\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}\s*$")

    NOMBRE_ABREVIADO = re.compile(r"(?i)^\s*nombre abreviado\s*$")

    AULA = re.compile(r"(?i)(aula|lab|sala|lb)")
    GRUPO_PRACTICO = re.compile(r"^\s*\(?\d{3}\)?\s*$")
    GRUPO_TEORICO = re.compile(r"^\s*\d{1,2}\s*$")

    CODIGO = re.compile(r"^\s*\d{4}-?E?\s*$")

    COL_CODIGO = re.compile(r"(?i)^\s*codigo\s*$")

    DIAS_SEMANA = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES"]


def clasificar_celda(valor_celda):
    if valor_celda is None:
        return ("VACIA", None)

    texto = str(valor_celda).strip()

    if ConfigurarHorario.HORA.match(texto):
        return ("COMIENZO_TABLA", texto)
    if ConfigurarHorario.NOMBRE_ABREVIADO.match(texto):
        return ("COMIENZO_TABLA_ASIGNATURAS", texto)
    if texto.upper() in ConfigurarHorario.DIAS_SEMANA:
        return ("DIA_SEMANA", texto.upper())

    if ConfigurarHorario.TRAMO.match(texto):
        return ("TRAMO_HORARIO", texto)
    
    if ConfigurarHorario.GRUPO_TEORICO.match(texto):
        return ("GRUPO_TEORICO", texto)

    if ConfigurarHorario.GRUPO_PRACTICO.match(texto):
        valor_limpio = texto.replace("(", "").replace(")", "")
        return ("GRUPO_PRACTICO", valor_limpio)

    if ConfigurarHorario.AULA.search(texto):
        return ("AULA", texto)

    if ConfigurarHorario.COL_CODIGO.match(texto):
        return ("COL_CODIGO", texto)

    if ConfigurarHorario.CODIGO.match(texto):
        return ("CODIGO_ASIG", texto)
        
    if texto.isupper() and texto not in ConfigurarHorario.DIAS_SEMANA:
        return ("ABREV_ASIG", texto)

    if not ConfigurarHorario.AULA.search(texto) and not ConfigurarHorario.GRUPO_TEORICO.match(texto) and not ConfigurarHorario.GRUPO_PRACTICO.match(texto) and not None:
        return ("ABREV_ASIG", texto)

    return ("TEXTO_GENERAL", texto)


# Creamos una función que crea el diccionario de celdas mergeadas O(n)
def celdas_mergeadas(ws):
    mapa_merge = {}
    for rango in ws.merged_cells.ranges:
        coord_origen = rango.start_cell.coordinate

        mapa_merge[coord_origen] = {
            "col_inicio": rango.min_col,
            "col_fin": rango.max_col,
            "fila_inicio": rango.min_row,
            "fila_fin": rango.max_row
        }

    return mapa_merge

def extraer_codigo_asignaturas(ws):
    diccionario_cod_asig = {}
    fila_inicio = 0
    columna_codigo = 0
    for fila in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            celda = ws.cell(row=fila, column=col)
            tipo, valor = clasificar_celda(celda.value)

            if tipo == "COL_CODIGO":
                columna_codigo = col
                fila_inicio = fila + 1
                break
        if columna_codigo != 0:
            break

    if columna_codigo != 0:
        for fila in range(fila_inicio, ws.max_row + 1):
            celda_cod = ws.cell(row=fila, column=columna_codigo)
            tipo_cod, valor_cod = clasificar_celda(celda_cod.value)
            if tipo_cod == "CODIGO_ASIG":
                celda_asig = ws.cell(row=fila, column=columna_codigo+2)
                tipo_asig, valor_asig = clasificar_celda(celda_asig.value)
    
                if tipo_asig == "ABREV_ASIG":
                    diccionario_cod_asig[valor_cod] = valor_asig
    return diccionario_cod_asig

def extraer_leyenda_asignaturas(ws, mapa_merge, diccionario_cod):
    diccionario_asignaturas = {}
    leyendo_tabla = False
    
    for fila in range(1, ws.max_row + 1):
        col = 1
        celda = ws.cell(row = fila, column = col)
        tipo, valor = clasificar_celda(celda.value)

        if tipo == "COMIENZO_TABLA_ASIGNATURAS":
            leyendo_tabla = True
            continue

        if leyendo_tabla:
            if valor == None or not str(valor).isupper():
                leyendo_tabla = False
                continue

            if celda.coordinate in mapa_merge:
                col_nombre_completo = mapa_merge[celda.coordinate]["col_fin"] + 1
            else:
                col_nombre_completo = col + 1

            celda_nombre_completo = ws.cell(row = fila, column = col_nombre_completo)

            if celda_nombre_completo.coordinate in mapa_merge:
                col_codigo = mapa_merge[celda_nombre_completo.coordinate]["col_fin"] + 1
            else:
                col_codigo = col_nombre_completo + 1

            celda_codigo = ws.cell(row=fila, column = col_codigo)
            tipo_c, valor_c = clasificar_celda(celda_codigo.value)

            if tipo_c == "CODIGO_ASIG":
                abrev_asig = diccionario_cod[valor_c]
            
                diccionario_asignaturas[abrev_asig] = valor_c
    return diccionario_asignaturas


def extraer_columnas_dias(ws, fila_inicio, mapa_merge):
    dias_rangos = {}

    for col in range(2, ws.max_column + 1):
        celda = ws.cell(row=fila_inicio, column=col)
        tipo, valor = clasificar_celda(celda.value)

        if tipo == "DIA_SEMANA":
            #print(valor)
            coord = celda.coordinate

            if coord in mapa_merge:
                limites = mapa_merge[coord]
                col_inicio = limites["col_inicio"]
                col_fin = limites["col_fin"]
            else:
                col_inicio = col
                col_fin = col

            dias_rangos[valor] = {
                "col_inicio": col_inicio,
                "col_fin": col_fin
            }
            
            if valor.upper() == "VIERNES":
                #print("Condición de parada de los días")
                break
                
    return dias_rangos


def extraer_filas_tramos(ws, fila_inicio, mapa_merge):
    tramos_horarios = []
    fila_actual = fila_inicio + 1
    contador = 0

    while fila_actual <= ws.max_row:
        celda = ws.cell(row=fila_actual, column = 1)
        tipo, valor = clasificar_celda(celda.value)

        if tipo == "TRAMO_HORARIO":
            coord = celda.coordinate

            if coord in mapa_merge:
                f_inicio = mapa_merge[coord]["fila_inicio"]
                f_fin = mapa_merge[coord]["fila_fin"]
            else:
                f_inicio = fila_actual
                f_fin = fila_actual

            tramos_horarios.append({
                "hora": valor,
                "fila_inicio": f_inicio,
                "fila_fin": f_fin
            })

            fila_actual = f_fin
        else:
            contador+=1
            if contador > 0:
                #print("Condición de parada tramos horarios")
                break
                
        fila_actual += 1
                
    return tramos_horarios


def calcular_horario_clase(lista_tramos, fila_inicio, fila_fin):
    hora_inicio = "00:00"
    hora_fin = "00:00"

    for tramo in lista_tramos:
        if tramo["fila_inicio"] <= fila_inicio <= tramo["fila_fin"]:
            hora_inicio = tramo["hora"].split("-")[0].strip()

        if tramo["fila_inicio"] <= fila_fin <= tramo["fila_fin"]:
            hora_fin = tramo["hora"].split("-")[1].strip()

    return hora_inicio, hora_fin

def calcular_dia(mapa_dias, col):
    dia = ""

    for d in mapa_dias:
        if mapa_dias[d]["col_inicio"] <= col <= mapa_dias[d]["col_fin"]:
            dia = d

    return dia

def extraer_aulas(valor):
    cont=0
    aulas={}
    for i, caracter in enumerate(valor):
        if caracter == "/":
            aulas[f"aula{cont}"] = valor.split("/")[cont].strip()
            cont+=1
            aulas[f"aula{cont}"] = valor.split("/")[cont].strip()

    if cont > 0:
        cond = True
    else:
        cond =  False
    return cond, cont + 1, aulas
            



def extraer_teoricas(ws, fila_inicio, fila_fin, col_actual, mapa_merge, celdas_visitadas, tabla, asignaturas, clases):
    dia = calcular_dia(tabla["mapa_dias"], col_actual)
    col_salto = col_actual + 1
    for f in range(fila_inicio, fila_fin + 1):
        valor_aula = "Aula 0"
        valor_a = "Aula 0"
        cond = False
        celda_asig = ws.cell(row=f, column=col_actual)
        tipo_asig, valor_asig = clasificar_celda(celda_asig.value)
        
        if tipo_asig != "ABREV_ASIG" or celda_asig.coordinate in celdas_visitadas:
            continue
            
        #print(f"\n[CLASE TEÓRICA] Asignatura: {valor_asig}, {asignaturas[valor_asig]}")
        celdas_visitadas.add(celda_asig.coordinate)

        fila_ini_clase = celda_asig.row
        
        if celda_asig.coordinate in mapa_merge:
            col_aula = mapa_merge[celda_asig.coordinate]["col_fin"] + 1
            fila_fin_clase = mapa_merge[celda_asig.coordinate]["fila_fin"]
        else:
            col_aula = col_actual + 1
            fila_fin_clase = fila_ini_clase

        hora_inicio, hora_fin = calcular_horario_clase(tabla["lista_tramos"], fila_ini_clase, fila_fin_clase)
        #print(f"Día: {dia}")
        #print(f"Horario: {hora_inicio} - {hora_fin}")

        celda_aula = ws.cell(row=f, column=col_aula)
        tipo_aula, valor_aula = clasificar_celda(celda_aula.value)
        
        if tipo_aula == "AULA":
            cond, num_aulas, aulas = extraer_aulas(valor_aula)
            #print(f"AULA: {valor_aula}, color: {celda_aula.fill.start_color.theme}")
            celdas_visitadas.add(celda_aula.coordinate)
            
            if celda_aula.coordinate in mapa_merge:      
                col_grupo = mapa_merge[celda_aula.coordinate]["col_fin"] + 1
            else:
                col_grupo = col_aula + 1
                
            celda_grupo = ws.cell(row=f, column=col_grupo)
            tipo_grupo, valor_grupo = clasificar_celda(celda_grupo.value)
            
            if tipo_grupo == "GRUPO_TEORICO":
                #print(f"Grupo teórico: {valor_grupo}")
                celdas_visitadas.add(celda_grupo.coordinate)

                if cond:
                    for aula in aulas.values():
                        nueva_clase = Clase(asig=valor_asig, cod_asig=asignaturas[valor_asig], grupo=valor_grupo, aula=aula, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="T")
                        clases.append(nueva_clase)
                else:
                    nueva_clase = Clase(asig=valor_asig, cod_asig=asignaturas[valor_asig], grupo=valor_grupo, aula=valor_aula, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="T")
                    clases.append(nueva_clase)
                
                if celda_grupo.coordinate in mapa_merge:
                    nuevo_salto = mapa_merge[celda_grupo.coordinate]["col_fin"] + 1
                else:
                    nuevo_salto = col_grupo + 1
                    
                if nuevo_salto > col_salto:
                    col_salto = nuevo_salto 
            else:
                if cond:
                    for aula in aulas.values():
                        if celda_asig.fill.start_color.theme == 9:
                            valor_grupo = "1"
                            nueva_clase = Clase(asig=valor_asig, cod_asig=asignaturas[valor_asig], grupo=valor_grupo, aula=aula, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="T")
                            clases.append(nueva_clase)
                        elif celda_asig.fill.start_color.theme == 7:
                            valor_grupo = "80"
                            nueva_clase = Clase(asig=valor_asig, cod_asig=asignaturas[valor_asig], grupo=valor_grupo, aula=aula, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="T")
                            clases.append(nueva_clase)
                else:
                    if celda_asig.fill.start_color.theme == 9:
                        valor_grupo = "1"
                        #print(f"Grupo teórico: {valor_grupo}")
                        nueva_clase = Clase(asig=valor_asig, cod_asig=asignaturas[valor_asig], grupo=valor_grupo, aula=valor_aula, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="T")
                        clases.append(nueva_clase)
                    elif celda_asig.fill.start_color.theme == 7:
                        valor_grupo = "80"
                        #print(f"Grupo teórico: {valor_grupo}")
                        nueva_clase = Clase(asig=valor_asig, cod_asig=asignaturas[valor_asig], grupo=valor_grupo, aula=valor_aula, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="T")
                        clases.append(nueva_clase)

        else:
            celda_abajo = ws.cell(row=f + 1, column=col_actual)
            tipo_a, valor_a = clasificar_celda(celda_abajo.value)
            if tipo_a == "AULA":
                cond, num_aulas, aulas = extraer_aulas(valor_a)
                celdas_visitadas.add(celda_abajo.coordinate)
            else:
                valor_a = "Aula 0"
                cond = False

            if celda_asig.fill.start_color.theme == 7:
                valor_grupo = "80"
                if cond:
                    for aula in aulas.values():
                        nueva_clase = Clase(asig=valor_asig, cod_asig=asignaturas[valor_asig], grupo=valor_grupo, aula=aula, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="T")
                        clases.append(nueva_clase)
                else:
                    nueva_clase = Clase(asig=valor_asig, cod_asig=asignaturas[valor_asig], grupo=valor_grupo, aula=valor_a, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="T")
                    clases.append(nueva_clase)
            elif celda_asig.fill.start_color.theme == 9:
                valor_grupo = "1"
                if cond:
                    for aula in aulas.values():
                        nueva_clase = Clase(asig=valor_asig, cod_asig=asignaturas[valor_asig], grupo=valor_grupo, aula=aula, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="T")
                        clases.append(nueva_clase)
                else:
                    nueva_clase = Clase(asig=valor_asig, cod_asig=asignaturas[valor_asig], grupo=valor_grupo, aula=valor_a, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="T")
                    clases.append(nueva_clase)
                
    return col_salto
    

def extraer_clase_practica_grupo(ws, fila_actual, col_actual, mapa_merge, celdas_visitadas, tabla, asignaturas, clases):
    celda = ws.cell(row=fila_actual, column=col_actual)
    coord = celda.coordinate
    _, valor = clasificar_celda(celda.value) # 'valor' será el nombre del grupo
    
    dia = calcular_dia(tabla["mapa_dias"], col_actual)
    fila_ini_clase = celda.row

    celda_abajo = ws.cell(row=fila_actual + 1, column=col_actual)
    tipo_a, valor_a = clasificar_celda(celda_abajo.value)

    if tipo_a == "ABREV_ASIG":
        celdas_visitadas.add(celda_abajo.coordinate)
        
    if coord in mapa_merge:
        col_aula = mapa_merge[coord]["col_fin"] + 1
    else:
        col_aula = col_actual + 1
                                
    celda_derecha = ws.cell(row=fila_actual, column=col_aula)
    tipo_d, valor_d = clasificar_celda(celda_derecha.value)

    if celda_derecha.coordinate in mapa_merge:
        fila_fin_clase = mapa_merge[celda_derecha.coordinate]["fila_fin"]
        salto_col = mapa_merge[celda_derecha.coordinate]["col_fin"] + 1
    else:
        fila_fin_clase = fila_ini_clase
        salto_col = col_aula + 1

    hora_inicio, hora_fin = calcular_horario_clase(tabla["lista_tramos"], fila_ini_clase, fila_fin_clase)

    cond = False
    if tipo_d == "AULA":
        celdas_visitadas.add(celda_derecha.coordinate)
        cond, _, aulas = extraer_aulas(valor_d)
    elif valor_d is None:
        valor_d = "Aula 0"
        cond = False

    if cond:
        for aula in aulas.values():
            nueva_clase = Clase(asig=valor_a, cod_asig=asignaturas[valor_a], grupo=valor, aula=aula, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="P")
            clases.append(nueva_clase)
            #print(f"[CLASE PRÁCTICA] Asignatura: {valor_a}, {asignaturas[valor_a]}")
            #print(f"Día: {dia}")
            #print(f"Horario: {hora_inicio} - {hora_fin}")
            #print(f"Aula: {aula}")
            #print(f"Grupo práctico: {valor}")
            #print("")

    else:
        nueva_clase = Clase(asig=valor_a, cod_asig=asignaturas[valor_a], grupo=valor, aula=valor_d, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="P")
        clases.append(nueva_clase)
        #print(f"[CLASE PRÁCTICA] Asignatura: {valor_a}, {asignaturas[valor_a]}")
        #print(f"Día: {dia}")
        #print(f"Horario: {hora_inicio} - {hora_fin}")
        #print(f"Aula: {valor_d}")
        #print(f"Grupo práctico: {valor}")
        
    return salto_col

def extraer_clase_rotada90(ws, fila_actual, col_actual, mapa_merge, celdas_visitadas, tabla, asignaturas, clases):
    celda = ws.cell(row=fila_actual, column=col_actual)
    coord = celda.coordinate
    tipo, valor = clasificar_celda(celda.value)
    
    dia = calcular_dia(tabla["mapa_dias"], col_actual)
    fila_ini_clase = celda.row 
    
    if coord in mapa_merge:
        col_grupo = mapa_merge[coord]["col_fin"] + 1
        fila_fin_clase = mapa_merge[coord]["fila_fin"]
    else:
        col_grupo = col_actual + 1
        fila_fin_clase = fila_ini_clase

    hora_inicio, hora_fin = calcular_horario_clase(tabla["lista_tramos"], fila_ini_clase, fila_fin_clase)
        
    celda_grupo = ws.cell(row=fila_actual, column=col_grupo)
    tipo_d, valor_d = clasificar_celda(celda_grupo.value)

    if tipo_d == "GRUPO_PRACTICO":
        celdas_visitadas.add(celda_grupo.coordinate)
        if celda_grupo.coordinate in mapa_merge:
            col_aula = mapa_merge[celda_grupo.coordinate]["col_fin"] + 1
        else:
            col_aula = col_grupo + 1 
    if celda.fill.start_color.theme == 7 and tipo_d != "GRUPO_PRACTICO":
        valor_d = "80"
        tipo = "T"
        if tipo_d == "AULA":
            col_aula = col_grupo
        else:
            col_aula = col_grupo + 1
    elif celda.fill.start_color.theme == 9 and tipo_d != "GRUPO_PRACTICO":
        valor_d = "1"
        tipo = "T"
        if tipo_d == "AULA":
            col_aula = col_grupo
        else:
            col_aula = col_grupo + 1
    elif celda.fill.start_color.theme == 6 and tipo_d == "GRUPO_PRACTICO":
        valor_d = "801"
        tipo = "P"
        if tipo_d == "AULA":
            col_aula = col_grupo
        else:
            col_aula = col_grupo + 1
    elif celda.fill.start_color.theme == 8 and tipo_d == "GRUPO_PRACTICO":
        valor_d = "101"
        tipo = "P"
        if tipo_d == "AULA":
            col_aula = col_grupo
        else:
            col_aula = col_grupo + 1
    else:
        if tipo_d == "AULA":
            col_aula = col_grupo
        else:
            col_aula = col_grupo + 1
        tipo = "P"  

    celda_aula = ws.cell(row=fila_actual, column=col_aula)
    tipo_d2, valor_d2 = clasificar_celda(celda_aula.value)

    cond = False
    if tipo_d2 == "AULA":
        celdas_visitadas.add(celda_aula.coordinate)
        cond, num_aulas, aulas = extraer_aulas(valor_d2)
        
        if celda_aula.coordinate in mapa_merge:
            salto_col = mapa_merge[celda_aula.coordinate]["col_fin"] + 1
        else:
            salto_col = col_aula + 1
    else:
        if valor_d2 is None:
            valor_d2 = "Aula 0"
        salto_col = col_aula + 1 

    
    if cond:
        for aula in aulas.values():
            nueva_clase = Clase(asig=valor, cod_asig=asignaturas[valor], grupo=valor_d, aula=aula, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo=tipo)
            clases.append(nueva_clase)
            #print(f"[CLASE PRÁCTICA] Asignatura: {valor}, {asignaturas[valor]}")
            #print(f"Día: {dia}")
            #print(f"Horario: {hora_inicio }- {hora_fin}")
            #print(f"Aula: {aula}")
            #print(f"Grupo práctico: {valor_d}")
            #print("")
    else:
        nueva_clase = Clase(asig=valor, cod_asig=asignaturas[valor], grupo=valor_d, aula=valor_d2, hora_inicio=hora_inicio,hora_fin=hora_fin, dia=dia, tipo=tipo)
        clases.append(nueva_clase)
        #print(f"[CLASE PRÁCTICA] Asignatura: {valor}, {asignaturas[valor]}")
        #print(f"Día: {dia}")
        #print(f"Horario: {hora_inicio} - {hora_fin}")
        #print(f"Aula: {valor_d2}")
        #print(f"Grupo práctico: {valor_d}")
        
    return salto_col


def extraer_practica_asig_rotacion0(ws, fila_actual, col_actual, mapa_merge, celdas_visitadas, tabla, asignaturas, clases):
    celda = ws.cell(row=fila_actual, column=col_actual)
    coord = celda.coordinate
    _, valor = clasificar_celda(celda.value) # 'valor' es la abreviatura de la asignatura
    
    dia = calcular_dia(tabla["mapa_dias"], col_actual)
    fila_ini_clase = celda.row

    if coord in mapa_merge:
        fila_aula = mapa_merge[coord]["fila_fin"] + 1
    else:
        fila_aula = fila_actual + 1
        
    celda_abajo = ws.cell(row=fila_aula, column=col_actual)
    tipo_a, valor_a = clasificar_celda(celda_abajo.value)

    cond = False
    if tipo_a == "AULA":
        cond, _, aulas = extraer_aulas(valor_a)
        celdas_visitadas.add(celda_abajo.coordinate)
    elif valor_a is None:
        valor_a = "Aula 0"
        cond = False

    if celda_abajo.coordinate in mapa_merge:
        fila_fin_clase = mapa_merge[celda_abajo.coordinate]["fila_fin"]
    else:
        fila_fin_clase = fila_ini_clase
        
    hora_inicio, hora_fin = calcular_horario_clase(tabla["lista_tramos"], fila_ini_clase, fila_fin_clase)

    if coord in mapa_merge:
        col_grupo = mapa_merge[coord]["col_fin"] + 1
    else:
        col_grupo = col_actual + 1
        
    celda_derecha = ws.cell(row=fila_actual, column=col_grupo)
    tipo_d, valor_d = clasificar_celda(celda_derecha.value)

    if tipo_d == "GRUPO_PRACTICO":
        celdas_visitadas.add(celda_derecha.coordinate)
        if celda_derecha.coordinate in mapa_merge:
            salto_col = mapa_merge[celda_derecha.coordinate]["col_fin"] + 1
        else:
            salto_col = col_grupo + 1
    else:
        if celda.fill.start_color.theme == 6:
            valor_d = 801
        elif celda.fill.start_color.theme == 8:
            valor_d = 101 
        salto_col = col_grupo 
    
    if cond:
        for aula in aulas.values():
            nueva_clase = Clase(asig=valor, cod_asig=asignaturas[valor], grupo=valor_d, aula=aula, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="P")
            clases.append(nueva_clase)
            #print(f"[CLASE PRÁCTICA] Asignatura: {valor}, {asignaturas[valor]}")
            #print(f"Día: {dia}")
            #print(f"Horario: {hora_inicio} - {hora_fin}")
            #print(f"Aula: {aula}")
            #print(f"Grupo Práctico: {valor_d}, color: {celda_derecha.fill.start_color.theme}")
    else:
        nueva_clase =  Clase(asig=valor, cod_asig=asignaturas[valor], grupo=valor_d, aula=valor_a, hora_inicio=hora_inicio, hora_fin=hora_fin, dia=dia, tipo="P")
        clases.append(nueva_clase)
        #print(f"[CLASE PRÁCTICA] Asignatura: {valor}, {asignaturas[valor]}")
        #print(f"Día: {dia}")
        #print(f"Horario: {hora_inicio} - {hora_fin}")
        #print(f"Aula: {valor_a}")
        #print(f"Grupo Práctico: {valor_d}, color: {celda_derecha.fill.start_color.theme}")
        
    return salto_col


def extraer_clases(ws, tabla, mapa_merge, asignaturas):
    clases = []


    celdas_visitadas = set()

    for dia, limites in tabla["mapa_dias"].items():
        for tramo in tabla["lista_tramos"]:
            fila_actual = tramo["fila_inicio"]
            col = limites["col_inicio"]
            while col <= limites["col_fin"]:
                celda = ws.cell(row=fila_actual, column=col)
                coord = celda.coordinate
                siguiente_col = col + 1
                if coord in celdas_visitadas:
                    col = siguiente_col
                    continue
                    
                tipo, _ = clasificar_celda(celda.value)
                
                if tipo in ["GRUPO_PRACTICO", "GRUPO_TEORICO", "AULA", "ABREV_ASIG"]:
                    color_tema = celda.fill.start_color.theme
                    color_rgb = celda.fill.start_color.rgb

                    #print("")

                    if color_tema in [2,7,9] and tipo == "ABREV_ASIG" and celda.alignment.textRotation == 90:
                        siguiente_col = extraer_clase_rotada90(ws, fila_actual, col, mapa_merge, celdas_visitadas, tabla, asignaturas, clases)
                    
                    elif color_tema in [2,7,9] and tipo == "ABREV_ASIG" and coord in mapa_merge:
                        siguiente_col = extraer_teoricas(ws, tramo["fila_inicio"], tramo["fila_fin"], col, mapa_merge, celdas_visitadas, tabla, asignaturas, clases)
                            
                    elif (color_tema in [6,8] or color_rgb == "FFF1FA78") and tipo == "GRUPO_PRACTICO":
                        siguiente_col = extraer_clase_practica_grupo(ws, fila_actual, col, mapa_merge, celdas_visitadas, tabla, asignaturas, clases)
                            
                    elif (color_tema in [6,8] or color_rgb == "FFF1FA78") and tipo == "ABREV_ASIG" and celda.alignment.textRotation == 90:
                        siguiente_col = extraer_clase_rotada90(ws, fila_actual, col, mapa_merge, celdas_visitadas, tabla, asignaturas, clases)
                                
                    elif (color_tema in [6,8] or color_rgb == "FFF1FA78") and tipo == "ABREV_ASIG" and celda.alignment.textRotation == 0:
                        siguiente_col = extraer_practica_asig_rotacion0(ws, fila_actual, col, mapa_merge, celdas_visitadas, tabla, asignaturas, clases)
                col = siguiente_col
    return clases


"""
Función para escanear la hoja, que detecta las tablas de horarios y extrae las clases de cada una de ellas
"""
def escanear_hoja(ws, diccionario_cod):
    clases_hoja = []
    mapa_merge = celdas_mergeadas(ws)

    asignaturas = extraer_leyenda_asignaturas(ws, mapa_merge, diccionario_cod)

    #print(asignaturas)
    
    fila_actual = 1
    max_row = ws.max_row
    
    while fila_actual <= max_row:
        celda = ws.cell(row=fila_actual, column = 1)
        tipo, valor = clasificar_celda(celda.value)
        
        if tipo == "COMIENZO_TABLA":
            #print("=========================")
            #print("He encontrado una tabla")
            #print("=========================")
            mapa_dias = extraer_columnas_dias(ws, fila_actual, mapa_merge)
            lista_tramos = extraer_filas_tramos(ws, fila_actual, mapa_merge)
            
            tabla = {
                "mapa_dias": mapa_dias,
                "lista_tramos": lista_tramos
            }
            
            clases_tabla = extraer_clases(ws, tabla, mapa_merge, asignaturas)
            
            for clase in clases_tabla:
                clases_hoja.append(clase)
            
            if lista_tramos:
                fila_actual = lista_tramos[-1]["fila_fin"]
            
            #break
            
        fila_actual += 1
    
    return clases_hoja

def extraer_hojas_analizar(ws):
    fila_inicio = 2
    hojas = []
    col=1

    while fila_inicio <= ws.max_row:
        celda = ws.cell(row=fila_inicio, column=col)
        if celda.value == None:
            break
        hojas.append(celda.value)
        fila_inicio += 1
    return hojas


def parsear_horario_excel(fichero):
    wb = openpyxl.load_workbook(fichero, data_only=True)
    diccionario_asignaturas = extraer_codigo_asignaturas(wb["ASIGNATURAS"])
    hojas_recorrer = extraer_hojas_analizar(wb.worksheets[-1])
    clases = []
    for hoja in hojas_recorrer:
        ws = wb[hoja]
        clases_hoja = escanear_hoja(ws, diccionario_asignaturas)
        clases.extend(clases_hoja)
    return clases

        